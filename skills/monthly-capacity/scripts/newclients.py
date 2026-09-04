"""
Stage the new clients approved in section 2 into the workbook, so the month can
be rebuilt with their hours resolved.

This exists because a client with hours but no entry in the app is flagged in
the review but its HOURS ARE NOT IN THE BUILD. `build_month.build()` only emits
assignment rows for job codes that resolve to a known client; anything else goes
to `report['unresolved_clients']` and shows up as the `unresolved` line in the
section 1 reconciliation.

`writeback.write_month(..., new_clients=[...])` adds the Clients rows to the
sheet, and nothing else. Used on its own it produces a client sitting in the app
with zero assignments while the month stays short by exactly that client's
hours. August 2026 would have written 11.13 hours short across two clients.

So the loop is: get the teams and budgets from the operator, `stage()` them into
a local copy of the workbook, REBUILD, confirm `unresolved` is 0, then write
back.

Roster rows matter as much as the client row. Role comes from the roster, never
from the timesheet — without a roster entry the build falls back to each
person's primary role, which is not necessarily the role the operator named.

Client ids assigned here are local to the rebuilt workbook and need not match
the ids `write_month` later assigns in the live sheet: writeback matches
assignment rows to clients BY NAME, so the two are deliberately uncoupled.

No client or staff names live in this file. Everything is read from the sheet.
"""
import openpyxl

# Mirrors writeback.ASSIGN_COLS. The Assignments tab is written with its exact
# header schema — never add a column to it (Code.gs rebuilds the whole row from
# the header and blanks anything its record does not know about).
ASSIGN_COLS = ['assignment_id', 'client_id', 'client_name',
               'staff_id', 'staff_name', 'role', 'hours']


def _header(ws):
    return [(str(c.value).strip() if c.value is not None else '') for c in ws[1]]


def _next_id(ws, col, prefix, width):
    n = 0
    for r in range(2, ws.max_row + 1):
        s = str(ws.cell(r, col).value or '')
        if s.startswith(prefix) and s[len(prefix):].isdigit():
            n = max(n, int(s[len(prefix):]))
    while True:
        n += 1
        yield f'{prefix}{n:0{width}d}'


def stage(app_path, clients, out_path=None):
    """Append approved new clients and their rosters to the workbook.

    clients: a list of dicts. Required keys:
        name  — exactly as the job code resolves in the review's section 2
        team  — [(staff_name, role), ...] the operator named in section 2
    Any other key that matches a Clients column is written to it (budgets,
    status, priority, notes). Budgets left out default to blank, which is how a
    client with no budget yet is represented — do not invent one.

    Returns {name: client_id}. Raises on a duplicate client or on a
    (staff, role) pair that is not on the Staff tab, rather than guessing.
    """
    wb = openpyxl.load_workbook(app_path)
    cw, aw, sw = wb['Clients'], wb['Assignments'], wb['Staff']

    ahdr = _header(aw)
    if ahdr != ASSIGN_COLS:
        raise RuntimeError(f'Assignments schema changed: {ahdr}')

    chdr = _header(cw)
    if 'name' not in chdr or 'client_id' not in chdr:
        raise RuntimeError(f'Clients schema unexpected: {chdr}')

    existing = {str(cw.cell(r, chdr.index('name') + 1).value or '').strip()
                for r in range(2, cw.max_row + 1)}

    shdr = _header(sw)
    staff_ids = {}
    for r in range(2, sw.max_row + 1):
        nm = str(sw.cell(r, shdr.index('name') + 1).value or '').strip()
        rl = str(sw.cell(r, shdr.index('role') + 1).value or '').strip()
        if nm and rl:
            staff_ids[(nm, rl)] = str(sw.cell(r, shdr.index('staff_id') + 1).value or '')

    gen_c = _next_id(cw, chdr.index('client_id') + 1, 'C', 4)
    gen_a = _next_id(aw, 1, 'A', 4)
    out = {}

    for spec in clients:
        name = str(spec['name']).strip()
        if not name:
            raise ValueError('a client was given with no name')
        if name in existing:
            raise ValueError(f'{name!r} is already on the Clients tab — '
                             'it is not a new client')
        team = spec.get('team') or []
        if not team:
            raise ValueError(f'{name!r} has no team. Section 2 asks for the '
                             'team precisely so the roster can be written; '
                             'without it the build guesses roles.')

        missing = [t for t in team if (str(t[0]).strip(), str(t[1]).strip())
                   not in staff_ids]
        if missing:
            raise KeyError(f'{name!r}: not on the Staff tab as that role: '
                           + ', '.join(f'{a} / {b}' for a, b in missing))

        cid = next(gen_c)
        rec = {k: v for k, v in spec.items() if k not in ('name', 'team')}
        rec['name'] = name
        rec['client_id'] = cid
        unknown = [k for k in rec if k not in chdr]
        if unknown:
            raise KeyError(f'{name!r}: no such Clients column: {unknown}')
        cw.append([rec.get(h, '') for h in chdr])

        for staff_name, role in team:
            staff_name, role = str(staff_name).strip(), str(role).strip()
            aw.append([next(gen_a), cid, name,
                       staff_ids[(staff_name, role)], staff_name, role, 0])

        existing.add(name)
        out[name] = cid

    wb.save(out_path or app_path)
    return out
