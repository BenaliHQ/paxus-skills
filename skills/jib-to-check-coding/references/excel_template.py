"""
Reference template for JIB check coding Excel files.
Copy and adapt this for each new JIB. Replace VENDOR, INVOICE_NUM, MONTH, YEAR,
TOTAL, and the rows list with the actual data.

After saving, recalc formulas:
    python3 /mnt/skills/public/xlsx/scripts/recalc.py <output_file>
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# === EDIT THESE ===
VENDOR = "Operator Name Here"
INVOICE_NUM = "12345"
MONTH = "March 2026"
TOTAL = 1234.56
OUTPUT_FILENAME = "Operator_Check_Coding_March_2026.xlsx"

# Each tuple: (CATEGORY, DESCRIPTION, AMOUNT, CLASS)
# Categories should be one of: "Well Costs - IDC", "Equipment", "Lease Operating Expense"
# (Or "Prepaid Expenses" for new cash calls, if that account exists)
ROWS = [
    ("Well Costs - IDC", "AFE -10 IDC - example (Mar 2026 JIB)", 100.00, "Example Well 1H"),
    ("Equipment", "AFE -20 Equipment Beyond Wellhead (Mar 2026 JIB)", 50.00, "Example Well 1H"),
    ("Lease Operating Expense", "LOE - chemicals, supervision (Mar 2026 JIB)", 25.00, "Example Well 1H"),
]

# === BUILD ===
wb = Workbook()
ws = wb.active
ws.title = "Check Coding"

# Header info (rows 1-4)
ws['A1'] = 'Vendor:'
ws['B1'] = VENDOR
ws['A2'] = 'Invoice #:'
ws['B2'] = INVOICE_NUM
ws['A3'] = 'Accounting Month:'
ws['B3'] = MONTH
ws['A4'] = 'Check Total:'
ws['B4'] = TOTAL
ws['B4'].number_format = '$#,##0.00'

for r in range(1, 5):
    ws[f'A{r}'].font = Font(name='Arial', size=11, bold=True)
    ws[f'B{r}'].font = Font(name='Arial', size=11)

# Column headers (row 6) — Paxus deep purple
headers = ['CATEGORY', 'DESCRIPTION', 'AMOUNT', 'CLASS']
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=6, column=i, value=h)
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='682145')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Data rows
start_row = 7
thin = Side(border_style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, (cat, desc, amt, cls) in enumerate(ROWS):
    r = start_row + i
    c1 = ws.cell(row=r, column=1, value=cat); c1.border = border; c1.font = Font(name='Arial', size=11)
    c2 = ws.cell(row=r, column=2, value=desc); c2.border = border; c2.font = Font(name='Arial', size=11)
    c3 = ws.cell(row=r, column=3, value=amt); c3.border = border; c3.font = Font(name='Arial', size=11)
    c3.number_format = '$#,##0.00;($#,##0.00)'
    c4 = ws.cell(row=r, column=4, value=cls); c4.border = border; c4.font = Font(name='Arial', size=11)

# Total row using SUM formula
total_row = start_row + len(ROWS)
lbl = ws.cell(row=total_row, column=2, value='Total')
lbl.font = Font(name='Arial', size=11, bold=True)
lbl.alignment = Alignment(horizontal='right')
total_cell = ws.cell(row=total_row, column=3, value=f'=SUM(C{start_row}:C{total_row-1})')
total_cell.font = Font(name='Arial', size=11, bold=True)
total_cell.number_format = '$#,##0.00;($#,##0.00)'
total_cell.fill = PatternFill('solid', start_color='F7EDF3')
for col in range(1, 5):
    ws.cell(row=total_row, column=col).border = Border(
        top=Side(border_style='medium'),
        bottom=Side(border_style='medium')
    )

# Column widths
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 75
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 30
ws.row_dimensions[6].height = 22

# Save to outputs
wb.save(f'/mnt/user-data/outputs/{OUTPUT_FILENAME}')
print(f'Saved: /mnt/user-data/outputs/{OUTPUT_FILENAME}')
